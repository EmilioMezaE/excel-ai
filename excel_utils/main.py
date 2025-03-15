import os
import json
import re
import uuid
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import openpyxl
from pydantic import BaseModel
from openai import OpenAI

# Load API Key from environment variables
AZURE_OPENAI_API_KEY = os.getenv("GITHUB_TOKEN")
if not AZURE_OPENAI_API_KEY:
    raise ValueError("Set your GitHub PAT token as an environment variable (GITHUB_TOKEN).")

# Initialize OpenAI client using Azure
client = OpenAI(
    base_url="https://models.inference.ai.azure.com",
    api_key=AZURE_OPENAI_API_KEY
)

# Initialize FastAPI app and CORS middleware
app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Change "*" to your frontend domain when deploying
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------
# In-memory storage for user sessions
# Each session has:
#   "conversation": list of {"role": "user"/"assistant", "content": "..."}
#   "excel_file": optional path if generated
# ---------------------------
user_sessions = {}

# ---------------------------
# Request Models
# ---------------------------
class ChatRequest(BaseModel):
    user_id: str
    message: str

# We'll return either a question or a link
# to the generated file
class ChatResponse(BaseModel):
    reply: str
    done: bool = False
    file_link: str = ""

# ---------------------------
# Utility: Build the Prompt for GPT
# We'll have a system instruction that says:
#   "You are an AI that helps gather info to build an Excel file...
#   If you have enough info, respond with JSON containing done=true
#   and the structure to build the Excel. Otherwise, ask more questions."
# ---------------------------
SYSTEM_PROMPT = """
You are an AI that helps gather info to build an Excel file. 
You will receive a conversation so far. If you still need more info from the user, 
ask a question. If you have enough info to build the Excel file, respond with JSON 
in this format exactly:

{
  "done": true,
  "excel_template": {
    "template_name": "...",
    "columns": [...],
    "formulas": {...},
    "formatting": {...}
  }
}

If you still need more info, respond with JSON:

{
  "done": false,
  "question": "Your next question here"
}

Do not include any other keys besides 'done', 'excel_template', or 'question'. 
No markdown formatting. 
"""

# ---------------------------
# /chat endpoint
# 1) Store user message in conversation
# 2) Call GPT with system prompt + conversation
# 3) Parse GPT's JSON
# 4) If done=true, build Excel and return link
# 5) Otherwise, return question
# ---------------------------
@app.post("/chat")
def chat_with_ai(req: ChatRequest):
    session = user_sessions.get(req.user_id)
    if not session:
        # Initialize new session
        session = {
            "conversation": [
                {"role": "system", "content": SYSTEM_PROMPT}
            ],
            "excel_file": None
        }
        user_sessions[req.user_id] = session

    # Append user message
    session["conversation"].append({"role": "user", "content": req.message})

    # Build GPT messages
    gpt_messages = session["conversation"]

    # Call GPT
    response = client.chat.completions.create(
        model="gpt-4o",  # or your available model
        messages=gpt_messages,
        temperature=0.7,
        max_tokens=800,
        top_p=1
    )

    raw_output = response.choices[0].message.content.strip()
    # Attempt to parse JSON
    # If parse fails, we'll just treat it as question
    raw_output = re.sub(r'```(json)?|```', '', raw_output).strip()

    try:
        parsed = json.loads(raw_output)
    except json.JSONDecodeError:
        # GPT did not return valid JSON, treat entire output as a question
        parsed = {
            "done": False,
            "question": raw_output
        }

    # If done = false, it's a question
    if not parsed.get("done"):
        question = parsed.get("question", raw_output)
        # Append GPT's question to conversation
        session["conversation"].append({"role": "assistant", "content": question})
        return {
            "reply": question,
            "done": False,
            "file_link": ""
        }

    # If done = true, we have an excel_template
    excel_template = parsed.get("excel_template")
    if not excel_template:
        # If no excel_template, fallback
        question = "I tried to parse excel_template but couldn't find it. Could you clarify?"
        session["conversation"].append({"role": "assistant", "content": question})
        return {
            "reply": question,
            "done": False,
            "file_link": ""
        }

    # Build the file
    file_path = create_excel_file(excel_template)
    session["excel_file"] = file_path

    # We'll return a link to download the file from /download/<filename> 
    # or we can just return an absolute path. For simplicity:
    file_name = os.path.basename(file_path)

    # Append GPT's "done" to conversation
    session["conversation"].append({"role": "assistant", "content": "Excel file generated."})
    user_sessions[req.user_id] = session

    return {
        "reply": "I've generated your Excel file! Click the link to download.",
        "done": True,
        "file_link": f"/download/{file_name}"
    }

# ---------------------------
# /download/{file_name} to serve the Excel
# ---------------------------
@app.get("/download/{file_name}")
def download_excel(file_name: str):
    file_path = os.path.join(os.getcwd(), file_name)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path, filename=file_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------------------
# Create Excel File from Template
# ---------------------------
def create_excel_file(template_data):
    filename = template_data.get("template_name", "ExcelFile").replace(" ", "_") + "_" + str(uuid.uuid4()) + ".xlsx"
    filepath = os.path.join(os.getcwd(), filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template_data.get("template_name", "Sheet1")

    # Add headers
    headers = template_data.get("columns", [])
    if headers:
        ws.append(headers)

    # Bold headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Insert formulas
    formulas = template_data.get("formulas", {})
    for cell_addr, formula in formulas.items():
        ws[cell_addr] = formula

    # Additional formatting if needed
    # e.g. template_data["formatting"]

    wb.save(filepath)
    return filepath
